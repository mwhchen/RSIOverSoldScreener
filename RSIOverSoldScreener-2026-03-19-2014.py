import yfinance as yf
import pandas as pd
import json
import time
import requests
import os
import glob
from io import StringIO
from datetime import datetime, timedelta
from tqdm import tqdm
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

def calculate_rsi(data, window=14):
    """
    Calculates the Relative Strength Index (RSI) using Wilder's Smoothing Method.
    """
    delta = data.diff()
    gain = (delta.where(delta > 0, 0))
    loss = (-delta.where(delta < 0, 0))

    avg_gain = gain.ewm(alpha=1/window, min_periods=window, adjust=False).mean()
    avg_loss = loss.ewm(alpha=1/window, min_periods=window, adjust=False).mean()

    rs = avg_gain / avg_loss
    rsi = 100 - (100 / (1 + rs))
    
    return rsi

def get_sp500_tickers():
    """
    Scrapes S&P 500 tickers from Wikipedia and adds SPY/^GSPC.
    """
    print("Fetching current S&P 500 ticker list from Wikipedia...")
    try:
        url = 'https://en.wikipedia.org/wiki/List_of_S%26P_500_companies'
        headers = {"User-Agent": "Mozilla/5.0"}
        response = requests.get(url, headers=headers)
        html = StringIO(response.text)
        df = pd.read_html(html)[0]
        tickers = df['Symbol'].str.replace('.', '-', regex=False).tolist()

        for t in ['SPY', '^GSPC']:
            if t not in tickers:
                tickers.append(t)
        return tickers
    except Exception as e:
        print(f"Error fetching tickers: {e}")
        return []

def get_latest_file(pattern):
    """
    Finds the most recent file matching a glob pattern based on modification time.
    """
    files = glob.glob(pattern)
    if not files:
        return None
    return max(files, key=os.path.getmtime)

def cleanup_old_files(pattern, keep=5):
    """
    Removes older versions of files to keep the repository size manageable.
    """
    files = sorted(glob.glob(pattern), key=os.path.getmtime, reverse=True)
    for old_file in files[keep:]:
        try:
            os.remove(old_file)
            print(f"Cleaned up old file: {old_file}")
        except Exception as e:
            print(f"Error cleaning up {old_file}: {e}")

def load_existing_data():
    """
    Loads data from the most recent timestamped JSON file inside the SP500 folder.
    """
    pattern = os.path.join("SP500", "sp500_data_*.json")
    filename = get_latest_file(pattern)
    if filename and os.path.exists(filename):
        print(f"Loading existing data from {filename}")
        with open(filename, 'r') as f:
            return json.load(f)
    return {"metadata": {"last_updated": "1970-01-01", "last_analysis": "1970-01-01"}, "tickers": {}}

def process_ticker(ticker, start_date, end_date):
    """
    Downloads data and calculates RSI for a specific ticker.
    """
    try:
        fetch_start = start_date - timedelta(days=40)
        df = yf.download(ticker, start=fetch_start, end=end_date, progress=False, auto_adjust=False)
        
        if df.empty:
            return None

        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)

        prices = df['Close']
        df['RSI'] = calculate_rsi(prices)
        
        df = df[df.index >= start_date]
        
        history = {}
        for date, row in df.iterrows():
            date_str = date.strftime('%Y-%m-%d')
            price_val = row['Close']
            rsi_val = row['RSI']
            
            history[date_str] = {
                "Price": round(float(price_val), 2),
                "RSI": round(float(rsi_val), 2) if pd.notna(rsi_val) else None
            }
        return history
    except Exception as e:
        return None

def send_to_discord(message, file_path=None):
    """
    Sends a message and optionally a file to a Discord Webhook.
    """
    webhook_url = os.getenv("DISCORD_WEBHOOK")
    
    if not webhook_url:
        print("Environment variable 'DISCORD_WEBHOOK' not set. Skipping Discord.")
        return

    try:
        if file_path and os.path.exists(file_path):
            with open(file_path, 'rb') as f:
                response = requests.post(
                    webhook_url,
                    data={"content": message},
                    files={"file": (os.path.basename(file_path), f)}
                )
        else:
            response = requests.post(webhook_url, json={"content": message})
        
        if response.status_code not in [200, 204]:
            print(f"Discord error: {response.status_code} - {response.text}")
    except Exception as e:
        print(f"Error sending to Discord: {e}")

def export_to_excel(df_results, ticker_summaries, global_summary, filename="rsi_recovery_analysis.xlsx"):
    """
    Handles the Excel file creation with refined formatting for the historical report.
    Includes columns for Day-3 to Day 0 RSI values. Summary block includes Sample Size.
    """
    try:
        header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        center_align = Alignment(horizontal="center")
        font_ticker = Font(size=20, bold=True)
        font_main = Font(size=16)
        medium_side = Side(border_style="medium", color="000000")
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            sheet_name = "Detailed Analysis"
            start_row = 0
            unique_tickers = df_results['Ticker'].unique()
            
            for ticker in tqdm(unique_tickers, desc="Formatting Excel Sheets"):
                pd.Series([f"TICKER: {ticker}"]).to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False, header=False)
                ws = writer.sheets[sheet_name]
                ws.cell(row=start_row + 1, column=1).font = font_ticker
                
                start_row += 1
                ticker_events = df_results[df_results['Ticker'] == ticker].drop(columns=['Ticker'])
                ticker_events.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
                
                num_cols = len(ticker_events.columns)
                for r in range(start_row + 1, start_row + len(ticker_events) + 2):
                    for c in range(1, num_cols + 1):
                        cell = ws.cell(row=r, column=c)
                        cell.font = font_main
                        if r == start_row + 1:
                            cell.fill = header_fill
                            cell.alignment = center_align
                        
                        # Apply borders to separate RSI data from Return data
                        # Columns: Date (1), RSI-3 (2), RSI-2 (3), RSI-1 (4), RSI-0 (5), Returns...
                        left_s = medium_side if c == 2 else None
                        right_s = medium_side if c == 5 else None
                        cell.border = Border(left=left_s, right=right_s)

                start_row += len(ticker_events) + 1
                summary_rows = ticker_summaries[ticker_summaries['Ticker'] == ticker].drop(columns=['Ticker'])
                
                # Align summary to start at Column E
                start_col_summary = 5
                summary_rows.to_excel(writer, sheet_name=sheet_name, startrow=start_row, startcol=start_col_summary - 1, index=False, header=False)
                
                for row_offset in range(len(summary_rows)):
                    for col_offset in range(len(summary_rows.columns)):
                        col_idx = start_col_summary + col_offset 
                        cell = ws.cell(row=start_row + 1 + row_offset, column=col_idx)
                        cell.font = font_main
                        
                        # Borders for the summary block
                        t_side = medium_side if row_offset == 0 else None
                        b_side = medium_side if row_offset == len(summary_rows) - 1 else None
                        r_side = medium_side if col_idx == start_col_summary else None
                        cell.border = Border(top=t_side, bottom=b_side, right=r_side)

                start_row += len(summary_rows) + 5

            ws = writer.sheets[sheet_name]
            for col in range(1, num_cols + 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

            ws_global = writer.book.create_sheet("Global Summary")
            ws_global.cell(row=1, column=1, value="GLOBAL TOTAL").font = font_ticker
            global_summary.to_excel(writer, sheet_name="Global Summary", startrow=1, index=False)
            
            for col_idx in range(1, len(global_summary.columns) + 1):
                cell = ws_global.cell(row=2, column=col_idx)
                cell.fill = header_fill
                cell.alignment = center_align
                cell.font = font_main
                ws_global.column_dimensions[ws_global.cell(row=2, column=col_idx).column_letter].width = 18
            
            for r in range(3, len(global_summary) + 3):
                for c in range(1, len(global_summary.columns) + 1):
                    ws_global.cell(row=r, column=c).font = font_main
            
        print(f"Excel report generated: {filename}")
        return filename
    except Exception as e:
        print(f"Error creating Excel file: {e}")
        return None

def export_current_signals_to_excel(df_signals, df_ref, filename):
    """
    Handles Excel output for current signals with styling.
    """
    try:
        header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        ref_header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        row_highlight_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        font_main = Font(size=14)
        font_header = Font(size=14, bold=True)
        font_label = Font(size=16, bold=True)
        font_label_unbold = Font(size=16, bold=False)
        font_alert = Font(size=14, bold=True, color="FF0000")
        font_legend = Font(size=12, italic=True, color="FF0000", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        hp_tickers = set()
        if not df_ref.empty:
            horizons = ["10d %", "30d %", "60d %", "90d %"]
            for ticker in df_ref['Ticker'].unique():
                ticker_rows = df_ref[df_ref['Ticker'] == ticker]
                avg_row = ticker_rows[ticker_rows['Metric'] == "Avg Rtn (%)"]
                hit_row = ticker_rows[ticker_rows['Metric'] == "Hit Rate (%)"]
                for h in horizons:
                    val_avg = avg_row[h].values[0] if not avg_row.empty else 0
                    val_hit = hit_row[h].values[0] if not hit_row.empty else 0
                    try:
                        if (isinstance(val_avg, (int, float)) and val_avg > 15) and (isinstance(val_hit, (int, float)) and val_hit > 75):
                            hp_tickers.add(ticker)
                            break
                    except: continue

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df_signals.to_excel(writer, sheet_name="Current Signals", startrow=5, index=False)
            ws = writer.sheets["Current Signals"]
            
            signal_date_val = datetime.now().strftime('%Y-%m-%d')
            ws.cell(row=1, column=1, value="Signal Date (Day 0)").font = font_label_unbold
            ws.cell(row=1, column=2, value=signal_date_val).font = font_label
            ws.cell(row=4, column=1, value="CURRENT SIGNALS").font = font_label
            legend_active = "LEGEND: RED/BOLD text indicates super high potential (Avg Return > 15% AND Hit Rate > 75%)"
            ws.cell(row=5, column=1, value=legend_active).font = font_legend
            
            for col in range(1, len(df_signals.columns) + 1):
                cell = ws.cell(row=6, column=col)
                cell.fill, cell.font, cell.alignment = header_fill, font_header, center_align
                ws.column_dimensions[ws.cell(row=6, column=col).column_letter].width = 22
            
            for r in range(7, len(df_signals) + 7):
                ticker_in_row = ws.cell(row=r, column=1).value
                is_high_perf = ticker_in_row in hp_tickers
                for c in range(1, len(df_signals.columns) + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.font, cell.alignment = (font_alert if is_high_perf else font_main), center_align

            if not df_ref.empty:
                start_row_title = len(df_signals) + 10
                ws.cell(row=start_row_title, column=1, value="HISTORICAL PERFORMANCE FOR ACTIVE TICKERS").font = font_label
                legend_hist = "LEGEND: Surpass Threshold (Avg Return > 15% OR Hit Rate > 75%)"
                ws.cell(row=start_row_title + 1, column=1, value=legend_hist).font = font_legend
                
                data_header_row = start_row_title + 2
                df_ref.to_excel(writer, sheet_name="Current Signals", startrow=data_header_row - 1, index=False)
                for col in range(1, len(df_ref.columns) + 1):
                    cell = ws.cell(row=data_header_row, column=col)
                    cell.fill, cell.font, cell.alignment, cell.border = ref_header_fill, font_header, center_align, thin_border

                unique_tickers = df_ref['Ticker'].unique()
                current_data_row = data_header_row + 1
                for i, ticker in enumerate(unique_tickers):
                    block_start, block_end = current_data_row, current_data_row + 2
                    ws.merge_cells(start_row=block_start, start_column=1, end_row=block_end, end_column=1)
                    ws.cell(row=block_start, column=1).alignment = center_align
                    current_fill = row_highlight_fill if i % 2 != 0 else None
                    for r in range(block_start, block_end + 1):
                        metric_type = ws.cell(row=r, column=2).value 
                        for c in range(1, len(df_ref.columns) + 1):
                            cell = ws.cell(row=r, column=c)
                            cell.border = thin_border
                            if current_fill: cell.fill = current_fill
                            cell.font, cell.alignment = font_main, center_align
                            if c > 2 and isinstance(cell.value, (int, float)):
                                if ("Avg Rtn" in str(metric_type) and cell.value > 15) or ("Hit Rate" in str(metric_type) and cell.value > 75):
                                    cell.font = font_alert
                    current_data_row += 3
        return hp_tickers 
    except Exception as e:
        print(f"Error creating Excel: {e}")
        return set()

def check_current_opportunities(data):
    """
    Checks for the RSI pattern and stores Excel in timestamped 'Scans' folder.
    """
    scans_folder = "Scans"
    if not os.path.exists(scans_folder): os.makedirs(scans_folder)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    file_path = os.path.join(scans_folder, f"current_rsi_signals_{timestamp}.xlsx")
    
    active_signals = []
    for ticker, history in tqdm(data["tickers"].items(), desc="Scanning"):
        sorted_dates = sorted(history.keys())
        if len(sorted_dates) < 4: continue
        d0, d1, d2, d3 = sorted_dates[-4:]
        r0, r1, r2, r3 = history[d0]["RSI"], history[d1]["RSI"], history[d2]["RSI"], history[d3]["RSI"]

        if r0 is not None and r0 <= 30 and r1 is not None and r1 > 30 and r3 is not None and r3 > 30:
            active_signals.append({
                "Ticker": ticker, "Day -3 RSI": r0, "Day -2 RSI": r1, "Day -1 RSI": r2, "Day 0 RSI": r3, "Current Price": history[d3]["Price"]
            })

    if not active_signals:
        send_to_discord(f"🚨 RSI Scan Complete: No new signals for {datetime.now().strftime('%Y-%m-%d')}.")
        return

    df_signals = pd.DataFrame(active_signals)
    ref_summary = []
    
    # Check inside RSI_Analysis folder for the report
    report_pattern = os.path.join("RSI_Analysis", "rsi_analysis_report_*.json")
    latest_report_file = get_latest_file(report_pattern)
    
    if latest_report_file:
        with open(latest_report_file, "r") as f:
            historical_data = json.load(f)
        horizons = ["10d %", "30d %", "60d %", "90d %"]
        for ticker in df_signals['Ticker'].unique():
            if ticker in historical_data:
                events = pd.DataFrame(historical_data[ticker])
                avg_row, hit_row, count_row = {"Ticker": ticker, "Metric": "Avg Rtn (%)"}, {"Ticker": ticker, "Metric": "Hit Rate (%)"}, {"Ticker": ticker, "Metric": "Sample Size"}
                for col in horizons:
                    valid = events[col].dropna()
                    if not valid.empty:
                        avg_row[col], hit_row[col], count_row[col] = round(valid.mean(), 2), round((valid > 0).sum()/len(valid)*100, 2), int(len(valid))
                    else:
                        avg_row[col], hit_row[col], count_row[col] = "N/A", "N/A", 0
                ref_summary.extend([avg_row, hit_row, count_row])
    
    df_ref = pd.DataFrame(ref_summary)
    hp_list = export_current_signals_to_excel(df_signals, df_ref, file_path)
    
    summary_msg = f"📈 **RSI Signals Found!** ({datetime.now().strftime('%Y-%m-%d')})\nTotal Tickers: {len(df_signals)}\n"
    if hp_list: summary_msg += f"🔥 **Super High Potential:** {', '.join(hp_list)}\n"
    send_to_discord(summary_msg, file_path)

def analyze_oversold_recovery(data, force_run=False):
    """
    Historical analysis with timestamped output stored in RSI_Analysis folder.
    Uses refined Excel export for historical summaries.
    """
    last_analysis_str = data.get("metadata", {}).get("last_analysis", "1970-01-01")
    days_since = (datetime.now() - datetime.strptime(last_analysis_str, '%Y-%m-%d')).days
    if not force_run and days_since < 90: return

    print("\nStarting Historical RSI Recovery Analysis...")
    analysis_results = {}
    all_events = []

    for ticker, history in tqdm(data["tickers"].items(), desc="Analyzing historical"):
        sorted_dates = sorted(history.keys())
        ticker_events = []
        for i in range(len(sorted_dates) - 4):
            # RSI pattern check over 4 days (d0 to d3)
            d0, d1, d2, d3, d4 = sorted_dates[i], sorted_dates[i+1], sorted_dates[i+2], sorted_dates[i+3], sorted_dates[i+4]
            r0, r1, r2, r3, p4 = history[d0]["RSI"], history[d1]["RSI"], history[d2]["RSI"], history[d3]["RSI"], history[d4]["Price"]
            
            # Pattern: RSI <= 30 at d0, then RSI > 30 at d1 and d3 (recovery)
            if r0 is not None and r0 <= 30 and r1 is not None and r1 > 30 and r3 is not None and r3 > 30:
                returns = {f"{h}d %": round(((history[sorted_dates[i+4+h]]["Price"] - p4) / p4) * 100, 2) if (i+4+h) < len(sorted_dates) else None for h in [10, 30, 60, 90]}
                event = {
                    "Ticker": ticker, 
                    "Day 0 Date": d3, 
                    "Day-3 RSI": round(r0, 2),
                    "Day-2 RSI": round(r1, 2) if r1 is not None else "N/A",
                    "Day-1 RSI": round(r2, 2) if r2 is not None else "N/A",
                    "Day 0 RSI": round(r3, 2) if r3 is not None else "N/A",
                    **returns
                }
                ticker_events.append(event)
                all_events.append(event)
        if ticker_events: analysis_results[ticker] = ticker_events

    data["metadata"]["last_analysis"] = datetime.now().strftime('%Y-%m-%d')
    
    df_results = pd.DataFrame(all_events)
    horizons = ["10d %", "30d %", "60d %", "90d %"]
    
    summary_list = []
    for ticker, events in analysis_results.items():
        df_t = pd.DataFrame(events)
        avg_rtn = {"Ticker": ticker, "Metric": "Avg Rtn (%)"}
        hit_rate = {"Ticker": ticker, "Metric": "Hit Rate (%)"}
        sample_size = {"Ticker": ticker, "Metric": "Sample Size"}
        for h in horizons:
            vals = df_t[h].dropna()
            avg_rtn[h] = round(vals.mean(), 2) if not vals.empty else "N/A"
            hit_rate[h] = round((vals > 0).sum() / len(vals) * 100, 2) if not vals.empty else "N/A"
            sample_size[h] = len(vals)
        summary_list.extend([avg_rtn, hit_rate, sample_size])
    
    ticker_summaries = pd.DataFrame(summary_list)
    
    global_avg = {"Metric": "Global Avg Rtn (%)"}
    global_hit = {"Metric": "Global Hit Rate (%)"}
    for h in horizons:
        vals = df_results[h].dropna()
        global_avg[h] = round(vals.mean(), 2) if not vals.empty else 0
        global_hit[h] = round((vals > 0).sum() / len(vals) * 100, 2) if not vals.empty else 0
    global_summary = pd.DataFrame([global_avg, global_hit])

    analysis_folder = "RSI_Analysis"
    if not os.path.exists(analysis_folder): os.makedirs(analysis_folder)
    
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    report_name = f"rsi_analysis_report_{timestamp}.json"
    report_path = os.path.join(analysis_folder, report_name)
    
    with open(report_path, "w") as f: json.dump(analysis_results, f, indent=4)
    
    excel_filename = os.path.join(analysis_folder, f"full_historical_rsi_analysis_{timestamp}.xlsx")
    excel_report_path = export_to_excel(df_results, ticker_summaries, global_summary, excel_filename)
    
    if excel_report_path:
        send_to_discord(f"📊 Historical RSI Analysis Complete! ({datetime.now().strftime('%Y-%m-%d')})\nReport saved: {report_name}", excel_report_path)
    else:
        send_to_discord(f"📊 Historical Analysis JSON Updated: {report_name}")
    
    cleanup_old_files(os.path.join(analysis_folder, "rsi_analysis_report_*.json"))
    cleanup_old_files(os.path.join(analysis_folder, "full_historical_rsi_analysis_*.xlsx"))

def RSI_history():
    """
    Main execution to update historical data with dynamic timestamping stored in SP500 folder.
    """
    data = load_existing_data()
    end_date = datetime.now()
    ten_years_ago = end_date - timedelta(days=365 * 10)
    current_sp500 = get_sp500_tickers()

    new_tickers = [t for t in current_sp500 if t not in data["tickers"]]
    existing_tickers = [t for t in current_sp500 if t in data["tickers"]]
    
    tickers_added = False
    if new_tickers:
        tickers_added = True
        for ticker in tqdm(new_tickers, desc="New tickers"):
            h = process_ticker(ticker, ten_years_ago, end_date)
            if h: data["tickers"][ticker] = h
            time.sleep(0.05)

    for ticker in tqdm(existing_tickers, desc="Updating tickers"):
        ticker_data = data["tickers"][ticker]
        dates = sorted(ticker_data.keys())
        latest = datetime.strptime(dates[-1], '%Y-%m-%d') if dates else ten_years_ago
        if (end_date - latest).days > 1:
            upd = process_ticker(ticker, latest, end_date)
            if upd: ticker_data.update(upd)

    data["metadata"]["last_updated"] = end_date.strftime('%Y-%m-%d')
    
    sp500_folder = "SP500"
    if not os.path.exists(sp500_folder): os.makedirs(sp500_folder)
    
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    filename = f"sp500_data_{timestamp}.json"
    file_path = os.path.join(sp500_folder, filename)
    
    with open(file_path, 'w') as f: json.dump(data, f, indent=4)
    send_to_discord(f"🔄 S&P 500 Database Updated: {filename}")
    cleanup_old_files(os.path.join(sp500_folder, "sp500_data_*.json"))
    return data, tickers_added

if __name__ == "__main__":
    current_data, new_tickers_found = RSI_history()
    analyze_oversold_recovery(current_data, force_run=new_tickers_found)
    check_current_opportunities(current_data)